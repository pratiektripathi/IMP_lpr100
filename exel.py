import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill


def create_report(cdata):

    today = datetime.datetime.today()
    xdate = today.strftime("%d-%m-%y")
    sno = 1
    party_data = {}
    for row in cdata:
        party = row[4]
        variety = row[5]
        if party not in party_data:
            party_data[party] = {}
        if variety not in party_data[party]:
            party_data[party][variety] = {'total_gross_wt': 0, 'total_tare_wt': 0, 'total_core_wt': 0, 'total_net_wt': 0, 'data': []}
            sno=1

        gross_wt, tare_wt, core_wt, net_wt = float(row[6]), float(row[7]), float(row[8]), float(row[9])
        party_data[party][variety]['total_gross_wt'] += gross_wt
        party_data[party][variety]['total_tare_wt'] += tare_wt
        party_data[party][variety]['total_core_wt'] += core_wt
        party_data[party][variety]['total_net_wt'] += net_wt

   
        del row[10:]
        del row[1:3]
        del row[2:4]
        
       
     
        row[0] = sno
        sno += 1
        party_data[party][variety]['data'].append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"

    title_font = Font(size=14, bold=True)
    center_alignment = Alignment(horizontal="center")

    # Fill color for the header row
    grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    ws['A1'] = "ADITYA FLEXIPACK PVT LTD"
    ws.merge_cells('A1:F1')
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    ws['A2'] = "Date:"
    ws['B2'] = today.strftime("%d-%m-%y")
    ws.merge_cells('B2:F2')


    row_index = 4

    for party, party_data in party_data.items():

        party_header = ["Party : ", party]
        ws.append(party_header)
        ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=6)
        for variety, variety_data in party_data.items():
            # Add the variety header
            variety_header = ["Job Name : " , variety]
            ws.append(variety_header)
            ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=6)
            # Apply grey fill to the header row


            # Add the table
            head = ["SNo.", "Roll No.", "Gross Wt.", "Tare Wt.", "Core Wt.", "Net Wt."]
            ws.append(head)
            for cell in ws[ws.max_row]:
                cell.fill = grey_fill

            table_data = variety_data['data']
            for item in table_data:
                ws.append(item)

            # Add the total row for this party and variety
            total_row = ['', 'Total', round(variety_data['total_gross_wt'], 2), round(variety_data['total_tare_wt'], 2),
                         round(variety_data['total_core_wt'], 2), round(variety_data['total_net_wt'], 2)]
            ws.append(total_row)

    wb.save("packing_list.xlsx")
